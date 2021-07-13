import openpyxl

class Testdatainput:
    @staticmethod
    def testdatainputexcel(tcname):
        Dict ={}

        book = openpyxl.load_workbook("C:\\Users\\Punam\\workspace_python\\ShopandFormpractice\\TestData\\Shopandform_data.xlsx")
        sheet = book.active

        for i in range(1, sheet.max_row+1):
            if sheet.cell(row=i, column=1).value == tcname:
                for j in range(2, sheet.max_column+1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [Dict]

